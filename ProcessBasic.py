import pandas as pd
import os 
import shutil
import openpyxl
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

# 1. 資料夾路徑相關

def create_folder(folder_name):
    """建立資料夾"""
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)
    return os.path.abspath(folder_name)

def delete_folders(deletelist):
    """
    刪除資料夾
    deletelist(list):需要為皆為路徑的list
    """
    for folder_name in deletelist: 
        if os.path.exists(folder_name): # 檢查資料夾是否存在
            shutil.rmtree(folder_name) # 刪除資料夾及其內容
        else:
            print(f"資料夾 '{folder_name}' 不存在。")

def find_folder(folderpath, find_by):
    """
    在指定的資料夾中尋找名稱包含特定字串的資料夾。
    Args:
        folderpath (str): 要搜尋的根目錄。
        find_by (str): 要匹配的字串，資料夾名稱中需包含此字串。
    Returns:
        list: 符合條件的資料夾完整路徑清單。
    """
    matching_folders = []
    
    for root, dirs, _ in os.walk(folderpath):
        for d in dirs:
            if find_by in d:
                matching_folders.append(os.path.join(root, d))
    return matching_folders

def check_pathexist(path):
    return os.path.exists(path)

def findfiles(filefolderpath, filetype='.csv', recursive=True):
    """
    尋找指定路徑下指定類型的檔案，並返回檔案路徑列表。

    Args:
        filefolderpath (str): 指定的檔案路徑。
        filetype (str, optional): 要尋找的檔案類型，預設為 '.csv'。
        recursive (bool, optional): 是否檢索所有子資料夾，預設為 True；反之為False，僅查找當前資料夾的所有file。

    Returns:
        list: 包含所有符合條件的檔案路徑的列表。
    """
    filelist = []

    if recursive:
        # 遍歷資料夾及其子資料夾
        for root, _, files in os.walk(filefolderpath):
            for file in files:
                if file.endswith(filetype):
                    file_path = os.path.join(root, file)
                    filelist.append(file_path)
    else:
        # 僅檢索當前資料夾
        for file in os.listdir(filefolderpath):
            file_path = os.path.join(filefolderpath, file)
            if os.path.isfile(file_path) and file.endswith(filetype):
                filelist.append(file_path)

    return filelist

def get_filename(path, extension=False):
    """
    從檔案路徑中提取檔名，可選擇是否包含副檔名。

    Args:
        path (str): 檔案的完整路徑。
        extension (bool, optional): 是否包含副檔名，預設為 False。

    Returns:
        str: 檔名（根據 extension 參數決定是否包含副檔名）。
    """
    filename = os.path.basename(path)
    if not extension:
        filename = os.path.splitext(filename)[0]
    return filename

def filter_basename(pathlist, strlist):
    return [path for path in pathlist if any(keyword in os.path.basename(path) for keyword in strlist)]

def getdatelist(time1, time2):
    '''
    建立日期清單
    time1、time2(str):為%Y-%M-%D格式的日期字串
    '''
    if time1 > time2:
        starttime = time2
        endtime = time1
    else:
        starttime = time1
        endtime = time2

    date_range = pd.date_range(start=starttime, end=endtime)
    datelist = [d.strftime("%Y%m%d") for d in date_range]
    return datelist

def copyfile(originalpath, newpath=None):
    """複製檔案，並且把檔案加上複製時間，並且回傳檔案路徑"""
    try:
        if not os.path.exists(originalpath):
            print("找不到原始檔案，請確認路徑是否正確。")
            return
        
        # 如果沒指定 newpath，自動產生帶日期的複製檔名
        if newpath is None:
            dirname, filename = os.path.split(originalpath)
            name, ext = os.path.splitext(filename)
            date_str = datetime.now().strftime('%Y%m%d')
            new_filename = f"{name}_{date_str}複製{ext}"
            newpath = os.path.join(dirname, new_filename)
        
        shutil.copy(originalpath, newpath)
        # print(f"檔案已成功複製到: {newpath}")
    except PermissionError:
        print("沒有權限複製檔案，請檢查權限設定。")
    except Exception as e:
        print(f"發生錯誤: {e}")

    return newpath

def movefile(originalpath, desfolder):
    """
    將檔案從原始路徑移動到指定資料夾。

    Args:
        originalpath (str): 檔案的原始路徑 (包含檔名)。
        desfolder (str): 目標資料夾路徑。
    """
    # 確保目標資料夾存在
    os.makedirs(desfolder, exist_ok=True)
    # 提取檔名
    filename = os.path.basename(originalpath)
    # 建立目標檔案路徑
    despath = os.path.join(desfolder, filename)
    # 移動檔案
    shutil.move(originalpath, despath)
    print(f"檔案已從 {originalpath} 移動至 {despath}")

def getfolderpath(path):
    '''返回當前該檔案資料夾位置'''
    # 檢查路徑是否有副檔名
    if os.path.isfile(path):
        # 如果是檔案，返回所在資料夾路徑
        return os.path.dirname(path)
    else:
        # 如果是資料夾，直接返回原本的路徑
        return path

def get_projectfolderpath(step=2):
    """
    根據當前工作目錄，找到 OneDrive 下指定層級的專案資料夾路徑。
    
    Args:
        step (int, optional): OneDrive 之後向下尋找的層級數，預設為 2，若要改onedrive資料夾則設為0。
    Returns:
        str: 指定層級的專案資料夾完整路徑。
    Raises:
        ValueError: 如果無法找到 OneDrive 目錄，則拋出錯誤。
    """
    current_path = Path(os.getcwd())
    
    # 找到 OneDrive 目錄名稱（支援「OneDrive - 公司名稱」）
    parts = current_path.parts
    try:
        onedrive_index = next(i for i, part in enumerate(parts) if "OneDrive" in part)
    except StopIteration:
        raise ValueError("OneDrive 資料夾未在當前工作目錄中找到")

    # 確保不超出可用範圍
    target_index = min(onedrive_index + step, len(parts) - 1)
    project_folder = Path(*parts[:target_index + 1])

    return str(project_folder)

def get_filename_withoutprojectname(path, step = 0):
    """
    從檔案路徑中提取不包含專案名稱的檔名。
    
    Args:
        path (str): 檔案的完整路徑。
    
    Returns:
        str: 移除專案資料夾後的檔案名稱。
    """
    project_folder = get_projectfolderpath(step)
    relative_path = path.replace(project_folder, "~")
    return relative_path

# 2. Excel 資料處理相關

def read_combined_dataframe(file_list):
    dataframes = []
    
    for file in file_list:
        try:
            if file.endswith('.csv'):
                df = pd.read_csv(file)
            elif file.endswith('.shp'):
                df = gpd.read_file(file)
            elif file.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file)
            else:
                print(f"Unsupported file format: {file}")
                continue
                
            dataframes.append(df)
        except Exception as e:
            print(f"Error reading {file}: {e}")

    # 合併所有 DataFrame
    combined_df = pd.concat(dataframes, ignore_index=True)
    return combined_df

def move_column(df, column_name, insert_index):
    """
    移動DataFrame中的既存欄位到指定位置。

    Args:
        df (pd.DataFrame): 要操作的DataFrame。
        column_name (str): 要移動的欄位名稱。
        insert_index (int): 欲插入的新位置索引（從0開始）。

    Returns:
        pd.DataFrame: 調整後的DataFrame。
    """
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' does not exist in DataFrame.")
    
    columns = df.columns.tolist() # 取得目前欄位順序
    columns.remove(column_name) # 移除該欄位
    columns.insert(insert_index, column_name) # 在指定位置插入該欄位
    return df[columns] # 重新排列DataFrame
    
def get_excel_sheet_names(path):
    """
    取得 Excel 檔案中的所有工作表名稱。

    Args:
        path (str): Excel 檔案的路徑。

    Returns:
        list: 工作表名稱列表。
    """
    try:
        sheet_names = pd.ExcelFile(path).sheet_names
        return sheet_names
    except FileNotFoundError:
        print(f"檔案不存在：{path}")
        return []
    except Exception as e:
        print(f"發生錯誤：{e}")
        return []

def duplicate_excel_sheet(excelpath, originalsheet, duplicatesheet, verbose = False):
    """
    建立excel工作頁副本。

    Args:
        excelpath (str): excel路徑。
        originalsheet (str): 建立副本的原始工作頁。
        duplicatesheet(str): 副本工作頁名稱。
        verbose(Boolean): 是否印出文字。

    Returns:
        DataFrame: 包含新插入的 Percent 欄位的資料框。
    """
    # 載入 Excel 文件
    wb = openpyxl.load_workbook(excelpath)
    
    # 確認原始工作表是否存在
    if originalsheet not in wb.sheetnames:
        print(f"工作表 {originalsheet} 不存在！")
        return
    
    # 取得原始工作表
    original = wb[originalsheet]
    
    # 複製原始工作表
    copied_sheet = wb.copy_worksheet(original)
    copied_sheet.title = duplicatesheet  # 設定副本工作表名稱
    
    # 儲存文件
    wb.save(excelpath)
    wb.close()
    if verbose:
        print(f"工作表 {originalsheet} 已成功複製為 {duplicatesheet}！")

def clean_excel_data(file_path, sheet_name, start_col='B', start_row=2, axis='range', end_col=None, end_row=None, verbose=False):
    """
    清除 Excel 檔案中指定工作表的資料與公式，讓儲存格變成完全空白。

    Args:
        file_path (str): Excel 檔案路徑。
        sheet_name (str): 工作表名稱。
        start_col (str): 清除資料的起始欄 (如 'B')。
        start_row (int): 清除資料的起始列 (如 2)。
        axis (str): 清除的範圍，'row' 表示整列，'col' 表示整欄，'range' 表示指定範圍。
        end_col (str): 清除資料的結束欄 (如 'D')，僅在 axis='range' 時有效。
        end_row (int): 清除資料的結束列 (如 10)，僅在 axis='range' 時有效。
        verbose (bool): 是否印出清除範圍的訊息 (預設 False)。
    """
    # 打開 Excel 檔案
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # 將欄位轉換成索引
    start_col_index = openpyxl.utils.column_index_from_string(start_col)
    end_col_index = openpyxl.utils.column_index_from_string(end_col) if end_col else start_col_index

    # 清除資料與公式
    if axis == 'row':  # 清除整列
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=start_row, column=col).value = None
    elif axis == 'col':  # 清除整欄
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=start_col_index).value = None
    elif axis == 'range':  # 清除指定範圍
        end_row = end_row if end_row else sheet.max_row
        for row in range(start_row, end_row + 1):
            for col in range(start_col_index, end_col_index + 1):
                sheet.cell(row=row, column=col).value = None
    else:
        raise ValueError("axis 必須是 'row', 'col', 或 'range'")

    # 儲存檔案
    wb.save(file_path)

    if verbose:
        print(f"已清除 {sheet_name} 的 {start_col}{start_row} 到 {end_col or start_col}{end_row or sheet.max_row} 範圍的資料與公式！")

def write_to_excel(excelpath, sheetname, cell, value, verbose = False):
    """
    在指定 Excel 工作表的指定儲存格填入數值。

    Args:
        excelpath (str): Excel 檔案的路徑。
        sheetname (str): 目標工作表名稱。
        cell (str): 目標儲存格 (如 'G4', 'H4')。
        value: 要填入的數值。
        verbose(Bool) : 是的話則會印出完成指定字串。

    Returns:
        None
    """
    wb = openpyxl.load_workbook(excelpath)
    
    # 確保目標工作表存在
    if sheetname not in wb.sheetnames:
        print(f"⚠️ 錯誤：工作表 '{sheetname}' 不存在！")
        wb.close()
        return

    ws = wb[sheetname]
    ws[cell] = value  # 填入值
    
    # 儲存並關閉
    wb.save(excelpath)
    wb.close()
    if verbose:
        print(f"✅ 在 '{sheetname}' 的 {cell} 填入 '{value}'")

def clean_and_paste(excelpath, sheet_name, df, startcell, title=True, verbose = False):
    """
    清空指定 Excel 工作表的內容，並將 df 從 startcell 開始寫入。

    參數:
    - excelpath (str): Excel 檔案路徑
    - sheet_name (str): 要操作的工作表名稱
    - df (pd.DataFrame): 要寫入的 DataFrame
    - startcell (str): 例如 'A1'，表示從哪個儲存格開始寫入
    - title (bool): 是否寫入欄位名稱 (預設 True)
    - verbose(Bool) : 是的話則會印出完成指定字串。
    """
    # 讀取 Excel
    wb = load_workbook(excelpath)

    if sheet_name not in wb.sheetnames:
        print(f"工作表 '{sheet_name}' 不存在！")
        return
    
    ws = wb[sheet_name]

    # 清空工作表 (移除所有內容)
    ws.delete_rows(1, ws.max_row)

    # 解析 startcell (A1 轉換成 row=1, col=1)
    col_letter = ''.join(filter(str.isalpha, startcell))  # 提取字母部分 (欄)
    row_number = int(''.join(filter(str.isdigit, startcell)))  # 提取數字部分 (列)
    start_col = column_index_from_string(col_letter)  # 轉換 A → 1, B → 2

    # 如果 title=True，先寫入欄位名稱
    if title:
        for c_idx, col_name in enumerate(df.columns, start=start_col):
            ws.cell(row=row_number, column=c_idx, value=col_name)
        row_number += 1  # 資料從下一列開始

    # 使用 openpyxl 方式將 DataFrame 寫入 (避免影響其他工作表)
    for r_idx, row in enumerate(df.to_numpy(), start=row_number):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # 儲存 Excel
    wb.save(excelpath)
    wb.close()
    if verbose:
        print(f"資料已寫入 {sheet_name}，起始位置：{startcell} (標題: {title})")

def paste_data_to_excel(file_path, sheet_name, data, start_col='B', start_row=2):
    """
    將資料貼到指定的 Excel 檔案與工作表中的固定欄位，保留其他公式。

    Args:
        file_path (str): Excel 檔案路徑。
        sheet_name (str): 工作表名稱。
        data (list): 要貼上的資料 (每個元素代表一列)。
        start_col (str): 貼上資料的起始欄 (如 'B')。
        start_row (int): 貼上資料的起始列 (預設從第 2 列開始)。
    """
    # 打開 Excel 檔案
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    # 將起始欄字母轉換為索引
    col_index = openpyxl.utils.column_index_from_string(start_col)

    # 貼資料到指定欄
    for i, value in enumerate(data, start=start_row):
        sheet.cell(row=i, column=col_index, value=value)

    # 儲存檔案
    wb.save(file_path)

def find_last_cell(excelpath, sheet_name=None):
    '''
    找到最後一筆資料在哪裡，返回"列"、"欄"

    Args:
        excelpath (str): Excel 檔案路徑。
        sheet_name (str, optional): 工作表名稱，沒有填寫的話會讀取第一個分頁。
    '''
    # 開啟 Excel 檔案
    workbook = openpyxl.load_workbook(excelpath, data_only=True)
    
    # 如果沒提供 sheet_name，預設使用第一個工作表
    sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    
    # 找最後有資料的列
    last_row = sheet.max_row
    while last_row > 0 and all(sheet.cell(row=last_row, column=col).value is None for col in range(1, sheet.max_column + 1)):
        last_row -= 1
    
    # 找最後有資料的欄
    last_column = sheet.max_column
    while last_column > 0 and all(sheet.cell(row=row, column=last_column).value is None for row in range(1, sheet.max_row + 1)):
        last_column -= 1
    
    # 把欄數轉成Excel字母
    last_column_letter = get_column_letter(last_column)
    
    workbook.close()
    return last_row, last_column_letter

def reformat_excel(excel_path, sheetname=None, allsheet=False, selectfont="微軟正黑體", fontsize=12):
    """自動調整列寬並設置字體格式
    Args:
        excel_path (str): 檔案的完整路徑。
        sheetname (str, optional) : 工作頁，若有要指定某一個工作頁使用。
        allsheet (bool, optional): 若要進行全部的工作業調整，請更改為True。
        selectfont (str) : 字體。
        fontsize (int) : 字體大小。
    """
    # 載入 Excel 文件
    wb = load_workbook(excel_path)

    # 根據是否選擇了特定工作表或處理所有工作表進行處理
    sheets_to_process = wb.sheetnames if allsheet else [sheetname] if sheetname else []

    if not sheets_to_process:
        # 如果沒有指定工作表名稱並且allsheet為False，則處理所有工作表
        sheets_to_process = wb.sheetnames

    for sheet in sheets_to_process:
        ws = wb[sheet]

        # 自動調整列寬
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 獲取列的字母名稱
            for cell in col:
                try:
                    # 避免空白格錯誤，並計算最長文字長度
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass  # 如果 cell 是空的，跳過
            # 計算並設置列寬
            adjusted_width = (max_length + 2) * 1.3
            ws.column_dimensions[column].width = adjusted_width

        # 設置字體
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name=selectfont, size=fontsize)

    # 儲存 Excel 文件
    wb.save(excel_path)

def merge_column_data(excel_path, sheet_name, columns, start_row=2, replace=True):
    """
    合併 Excel 指定欄位中相鄰且內容相同的儲存格，並進行跨欄置中對齊。

    參數：
    - excel_path (str): Excel 檔案的路徑
    - sheet_name (str): 目標工作表名稱
    - columns (list): 要合併的欄位名稱列表
    - start_row (int): 從哪一行開始合併（預設為 2）
    - replace (bool): 是否覆蓋原始檔案 (True=覆蓋, False=另存新檔)
    """
    
    # 讀取 Excel
    wb = openpyxl.load_workbook(excel_path)
    
    # 確保 sheet 存在
    if sheet_name not in wb.sheetnames:
        print(f"錯誤：找不到工作表 '{sheet_name}'")
        return

    sheet = wb[sheet_name]

    # 獲取總行數
    max_row = sheet.max_row

    for col in columns:
        col_index = None
        # 找到對應欄位的索引
        for i, cell in enumerate(sheet[1], start=1):
            if cell.value == col:
                col_index = i
                break
        
        if col_index is None:
            print(f"找不到欄位 {col}")
            continue

        # 開始合併相同內容的儲存格
        merge_start = start_row  # 設定合併起點
        for row in range(start_row + 1, max_row + 2):  # 從 start_row+1 行開始比對
            current_value = sheet.cell(row=merge_start, column=col_index).value
            next_value = sheet.cell(row=row, column=col_index).value

            if current_value != next_value or row > max_row:
                if row - merge_start > 1:
                    sheet.merge_cells(start_row=merge_start, start_column=col_index, 
                                      end_row=row-1, end_column=col_index)
                    merged_cell = sheet.cell(row=merge_start, column=col_index)
                    merged_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                merge_start = row  # 更新起始行數

    # 儲存 Excel 檔案
    if replace:
        wb.save(excel_path)
        print(f"合併完成，原檔案已覆蓋：{excel_path}")
    else:
        new_excel_path = excel_path.replace(".xlsx", "_merged.xlsx")
        wb.save(new_excel_path)
        print(f"合併完成，已另存為：{new_excel_path}")


# 3. 系統操作文件

def updatelog(file, text):
    """將 text 追加寫入指定的 log 檔案，並加上當前時間"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # 取得當前時間
    log_entry = f"[{timestamp}] {text}"  # 格式化日誌內容
    with open(file, 'a', encoding='utf-8') as f:
        f.write(log_entry + '\n')

def is_expired(line, cutoff_date):
    """判斷該行的時間戳記是否超過 `cutoff_date`"""
    try:
        timestamp_str = line[1:20]  # 擷取 `[YYYY-MM-DD HH:MM:SS]`
        log_time = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
        return log_time < cutoff_date
    except ValueError:
        return False  # 解析錯誤則保留該行

def refreshlog(file, day=30):
    """僅檢查第一行的時間戳記，若超過 `day` 天才執行清理"""
    if not os.path.exists(file):
        return  # 檔案不存在，直接返回

    cutoff_date = datetime.now() - timedelta(days=day)  # 計算過期時間

    with open(file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    if not lines:
        return  # 檔案為空，直接返回

    # 解析第一行的時間戳記
    first_line = lines[0]
    if first_line.startswith('['):  # 確保這行有時間戳記
        try:
            timestamp_str = first_line[1:20]  # 擷取 `[YYYY-MM-DD HH:MM:SS]`
            first_log_time = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
            if first_log_time >= cutoff_date:
                return  # 如果第一行時間還在範圍內，直接跳出
        except ValueError:
            pass  # 解析失敗就忽略，繼續清理

    # 若第一行時間超過 `day` 天，則開始過濾所有行
    new_lines = [line for line in lines if not (line.startswith('[') and is_expired(line, cutoff_date))]

    # 重新寫入檔案
    with open(file, 'w', encoding='utf-8') as f:
        f.writelines(new_lines)

# 4. 鼎漢資料慣用處理

def get_percent_columns(df, columns='Trips'):
    """
    計算百分比欄位，並插入到指定的 columns 欄位後面。

    Args:
        df (DataFrame): 輸入的資料框。
        columns (str): 用來計算百分比的欄位名稱。

    Returns:
        DataFrame: 包含新插入的 Percent 欄位的資料框。
    """
    total_value = df[columns].sum()
    df['Percent'] = (df[columns] / total_value) * 100
    df['Percent'] = df['Percent'].round(2).astype(str) + "%"

    # 找到 columns 欄位的位置，將 Percent 插入在其後面
    col_index = df.columns.get_loc(columns) + 1
    cols = list(df.columns)
    cols.insert(col_index, cols.pop(cols.index('Percent')))
    df = df[cols]

    return df


def keepZH_tw(df, keepsuffixies='_Zh_tw', deletesuffixies='_En'):
    # 刪除包含 deletesuffixies 的欄位
    df = df.loc[:, ~df.columns.str.endswith(deletesuffixies)]
    
    # 修改欄位名稱：去掉 keepsuffixies 的後綴
    df.columns = [col.replace(keepsuffixies, '') if col.endswith(keepsuffixies) else col for col in df.columns]
    return df

def get_VL1(df, Vcolumn, VLimitcolumn):
    df['V/VL'] = df[Vcolumn] / df[VLimitcolumn]

    conditions = [
        df['V/VL'] < 0.2,
        (df['V/VL'] >= 0.2) & (df['V/VL'] < 0.4),
        (df['V/VL'] >= 0.4) & (df['V/VL'] < 0.6),
        (df['V/VL'] >= 0.6) & (df['V/VL'] < 0.8),
        (df['V/VL'] >= 0.8) & (df['V/VL'] < 0.9),
        df['V/VL'] >= 0.9
    ]

    values = [6, 5, 4, 3, 2, 1]

    df['LOS_VL1'] = np.select(conditions, values, default=np.nan)  # 預設 NaN 避免錯誤
    return df

def get_VL2(df, Vcolumn, VLimitcolumn):
    df['V/VL'] = df[Vcolumn] / df[VLimitcolumn]

    conditions = [
        df['V/VL'] < 0.2,
        (df['V/VL'] >= 0.2) & (df['V/VL'] < 0.4),
        (df['V/VL'] >= 0.4) & (df['V/VL'] < 0.5),
        (df['V/VL'] >= 0.5) & (df['V/VL'] < 0.6),
        (df['V/VL'] >= 0.6) & (df['V/VL'] < 0.8),
        df['V/VL'] >= 0.8
    ]

    values = ['F', 'E', 'D', 'C', 'B', 'A']

    df['LOS_VL2'] = np.select(conditions, values, default=None)  # 預設 NaN 避免錯誤
    return df

def get_LOS_VC(df, Vcolumn, Ccolumn):
    df['V/C'] = df[Vcolumn] / df[Ccolumn]

    conditions = [
        df['V/C'] <= 0.25,
        (df['V/C'] > 0.25) & (df['V/C'] <= 0.50),
        (df['V/C'] > 0.50) & (df['V/C'] <= 0.80),
        (df['V/C'] > 0.80) & (df['V/C'] <= 0.90),
        (df['V/C'] > 0.90) & (df['V/C'] <= 1.00),
        df['V/C'] > 1.00
    ]

    values = ['A', 'B', 'C', 'D', 'E', 'F']

    df['LOS_V/C'] = np.select(conditions, values, default=np.nan)  # 預設 NaN 避免錯誤
    return df

def matrixtable(df, from_columns, to_columns):
    """
    根據指定的 'from_columns' 和 'to_columns' 生成 OD 矩陣格式的表格。
    
    Parameters:
    df (DataFrame): 原始數據框。
    from_columns (str): 要從中提取 "O" 的欄位名稱。
    to_columns (str): 要從中提取 "D" 的欄位名稱。
    
    Returns:
    DataFrame: 轉換後的 OD 矩陣表格。
    """
    # 使用pivot進行轉換
    od_matrix = df.pivot_table(index=from_columns, columns=to_columns, values='Value', aggfunc='first')

    # 重設索引，去掉列名
    od_matrix.reset_index(inplace=True)
    od_matrix.columns.name = None

    # 生成 'OD' 欄位
    od_matrix['OD'] = od_matrix[from_columns]
    od_matrix = od_matrix.drop(from_columns, axis=1)

    # 調整列的順序
    od_matrix = od_matrix[['OD'] + list(od_matrix.columns[:-1])]

    return od_matrix


def get_peak_data(df, group_by, sum_by, hourcolumn):
    """
    取得指定資料欄位中的尖峰時段資料，並返回最大PCU值對應的資料。

    Args:
        df (DataFrame): 要處理的資料集。
        group_by (str): 用來分組的欄位名稱。
        sum_by (str): 用來求最大值的欄位名稱。
        hourcolumn (str): 代表小時的欄位名稱。

    Returns:
        DataFrame: 包含每組尖峰時段資料及其對應的最大PCU值的資料集。
    """

    # 取得每組的最大 PCU 值對應的索引
    idx = df.groupby(group_by)[sum_by].idxmax()
    # 取出尖峰時段資料
    peak_hour = df.loc[idx].copy()
    # 重新命名欄位
    peak_hour = peak_hour.rename(columns={sum_by: '尖峰小時PCU', hourcolumn: '尖峰時段'})
    return peak_hour.reset_index(drop=True)

def get_peak_AMPM(df, group_by, sum_by, hourcolumn):
    """
    取得指定資料欄位中的晨峰及昏峰資料，並返回最大PCU值對應的資料。

    Args:
        df (DataFrame): 要處理的資料集。
        group_by (str): 用來分組的欄位名稱。
        sum_by (str): 用來求最大值的欄位名稱。
        hourcolumn (str): 代表小時的欄位名稱。

    Returns:
        DataFrame: 包含每組尖峰時段資料及其對應的最大PCU值的資料集。
    """

    df_AM = df[df[hourcolumn] <= 12].reset_index(drop = True)
    df_PM = df[df[hourcolumn] > 12].reset_index(drop = True)
    df_AM_peak = get_peak_data(df = df_AM, group_by=group_by, sum_by=sum_by, hourcolumn=hourcolumn).rename(columns = {'尖峰小時PCU':'晨峰小時PCU'})
    df_PM_peak = get_peak_data(df = df_PM, group_by=group_by, sum_by=sum_by, hourcolumn=hourcolumn).rename(columns = {'尖峰小時PCU':'昏峰小時PCU'})
    return df_AM_peak, df_PM_peak


def get_peak_percent(df, group_by, sum_by, hourcolumn):

    """
    取得指定資料欄位中的尖峰時段資料，並計算尖峰小時比例返回最大PCU值對應的資料。
    須包含 "get_peak_data" 這個函數。

    Args:
        df (DataFrame): 要處理的資料集。
        group_by (str): 用來分組的欄位名稱。
        sum_by (str): 用來求最大值的欄位名稱。
        hourcolumn (str): 代表小時的欄位名稱。

    Returns:
        DataFrame: 包含每組尖峰時段資料及其對應的最大PCU值的資料集。
    """

    df_peakdata = get_peak_data(df, group_by=group_by, sum_by=sum_by, hourcolumn=hourcolumn)
    df_daily = df.groupby(group_by).agg({sum_by:'sum'}).reset_index()
    df_daily = pd.merge(df_peakdata, df_daily)
    df_daily['尖峰率'] = df_daily['尖峰小時PCU'] / df_daily['PCU']
    df_daily = df_daily.drop(columns = ['尖峰時段', sum_by])
    output = pd.merge(df, df_daily, on = group_by, how='left')
    # 刪除所有以 '_y' 結尾的欄位
    output = output.drop(columns=[col for col in output.columns if col.endswith('_y')])
    # 將所有以 '_x' 結尾的欄位名稱中的 '_x' 改為空字符，或者你也可以改為其他文字
    output.columns = [col.replace('_x', '') for col in output.columns]
    output = output.drop_duplicates()
    return output    

# ========== 以下可用，但仍須修正 =========


def seperate_mergecolumns(excelpath, sheetname=None, replace=True):
    """
    將檔案跨欄置中的資料填入相同值。

    Args:
        excelpath (str): 檔案的原始路徑 (包含檔名)。
        sheetname (str, optional): 工作頁，如果沒有填的話則是處理第一個工作頁。
        replace (Boolean, optional) : 是否覆蓋，反之為另存。 
    """
    # 開啟 Excel 檔案
    wb = openpyxl.load_workbook(excelpath)
    
    # 如果沒提供 sheetname，就選擇第一個工作表
    sheet = wb[sheetname] if sheetname else wb.active

    # 如果 replace=False，就複製一份新檔案
    if not replace:
        base, ext = os.path.splitext(excelpath)
        new_excelpath = f"{base}_seperated{ext}"
    else:
        new_excelpath = excelpath

    # 複製合併儲存格的範圍
    merged_cells = list(sheet.merged_cells.ranges)

    # 解開合併儲存格並填入資料
    for merged_cell in merged_cells:
        min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
        value = sheet.cell(min_row, min_col).value
        
        sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row, col, value)

    # 儲存檔案
    wb.save(new_excelpath)
    print(f"已處理跨欄置中，儲存至：{new_excelpath}")

def get_seperatedcolumns_df(excelpath, sheetname=None):

    seperate_mergecolumns(excelpath = excelpath, sheetname = sheetname, replace = False) # 產生一個暫存檔

    # 讀取我們的暫存檔
    base, ext = os.path.splitext(excelpath)
    new_excelpath = f"{base}_seperated{ext}"
    new_excelpath = os.path.abspath(new_excelpath)
    df = pd.read_excel(new_excelpath, sheet_name = sheetname)

    # 刪除暫存檔
    os.remove(new_excelpath)
    return df 
